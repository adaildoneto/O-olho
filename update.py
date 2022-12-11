from bs4 import BeautifulSoup as bs
import requests
import json
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import pandas as pd



path = "thundera.xlsx"
wb_obj = load_workbook(path) 
  
sheet_obj = wb_obj.active 
sheet_obj.title = "Olho de Thundera"
m_row = sheet_obj.max_row

lista = []

for k in range(1, m_row + 1): 
   cell_obj = sheet_obj.cell(row = k, column = 5)  
   lista.append(cell_obj.value)




sites = ['http://ac24horas.com', 'http://contilnetnoticias.com.br', 'http://correio68.com', 'https://agencia.ac.gov.br', 'https://nahoradanoticia.com.br',
'http://folhadoacre.com.br', 'http://yaconews.com', 'http://jornalopiniao.net', 'http://3dejulhonoticias.com.br', 'https://noticiadoacre.com.br',
'http://ecosdanoticia.net.br', 'http://agazetadoacre.com', 'http://www.acre.com.br', 'http://acreagora.com', 'https://acreinfoco.com', 'https://acjornal.com',
'http://oaltoacre.com', 'http://agazeta.net', 'http://noticiasdahora.com.br',  'http://oacreagora.com', 'https://www.noticiasdafronteira.com.br', 
'https://www.juruaonline.com.br', 'https://www.juruaemtempo.com.br', 'https://oestadoacre.com', 'https://oseringal.com', 'https://www.vozdonorte.com.br', 
'https://amazoniaagora.com.br', 'https://acreonline.net' , 'https://acriticadoacre.com.br','http://portalquinari.com.br',
'https://oacre.com.br',]

jnews = len (sites)

for i in range (jnews):
    turl = (sites[i] + '/wp-json/wp/v2/posts')
    response = requests.get(turl)

    data = response.text
    
    if response.status_code == 200:  
      print('site ' + turl + ' ok!')
      dados = json.loads(data)
      for j in dados:
         titulo = j['title']['rendered']
         descricao = j['content']['rendered']
         link = j['link']
         side = sites[i]
         
         try:
          img = j['jetpack_featured_media_url']
         except KeyError:
          img = 'Imagem não encontrada'        
         
         try:
            cdata = j['date']
            cdata1 = cdata[0:10]
            chora = cdata[11:20]
         except KeyError:
            cdata = 0

         #estruturando o conteudo dentro da celula
         pdata = (cdata1, chora, titulo, descricao, link, img, side)

                
         
         if (link not in lista):
            sheet_obj.append (pdata)
         
    else:
      print('site ' + turl + ' não habilitado para o json')

# Salvando o planilha
wb_obj.save('thundera.xlsx')

# lendo a planilha The Eye of Thundera
df1 = pd.read_excel('thundera.xlsx')

# ordenando a planilha por data e hora
ok = df1.sort_values(by=['Data', 'Hora'])
print ('Planinha Atualizada')

ok.to_json(r'thunder-ordenado.json')
print ('Criando Json')

ok.to_excel("thundera-ordenado.xlsx")
print ('Planinha Atualizada')
